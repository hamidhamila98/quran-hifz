#!/usr/bin/env python3
"""
Pipeline OCR Arabe + Traduction Français
========================================
1. PDF → Images (pdf2image)
2. Images → Texte arabe (Google Cloud Vision)
3. Texte arabe → Français (Gemini)
4. Export → JSON + Excel
"""

import os
import json
import time
import base64
import requests
from pathlib import Path

# === CONFIGURATION ===
# ⚠️ CHANGE TA CLÉ API APRÈS UTILISATION (tu l'as partagée publiquement!)
API_KEY = "AIzaSyAcFwrCFv4e2Zr79S_y-d640EBNhqGGq5k"

# Modèle Gemini à utiliser
GEMINI_MODEL = "gemini-2.5-pro-preview-05-06"  # ou "gemini-2.0-flash" pour plus rapide/moins cher


def install_dependencies():
    """Installe les dépendances nécessaires"""
    import subprocess
    packages = [
        "pdf2image",
        "Pillow", 
        "openpyxl",
        "google-generativeai",
    ]
    for pkg in packages:
        subprocess.run(["pip", "install", pkg, "--break-system-packages", "-q"])
    # Installer poppler pour pdf2image
    subprocess.run(["apt-get", "update", "-qq"], capture_output=True)
    subprocess.run(["apt-get", "install", "-y", "-qq", "poppler-utils"], capture_output=True)


def pdf_to_images(pdf_path, output_dir="pages"):
    """Convertit un PDF en images (une par page)"""
    from pdf2image import convert_from_path
    
    Path(output_dir).mkdir(exist_ok=True)
    print(f"📄 Conversion du PDF en images...")
    
    images = convert_from_path(pdf_path, dpi=300)
    image_paths = []
    
    for i, img in enumerate(images, 1):
        img_path = f"{output_dir}/page_{i:04d}.png"
        img.save(img_path, "PNG")
        image_paths.append(img_path)
        print(f"  ✓ Page {i}/{len(images)}")
    
    return image_paths


def ocr_with_vision_api(image_path, api_key):
    """OCR d'une image avec Google Cloud Vision API"""
    
    # Lire et encoder l'image en base64
    with open(image_path, "rb") as f:
        image_content = base64.b64encode(f.read()).decode("utf-8")
    
    url = f"https://vision.googleapis.com/v1/images:annotate?key={api_key}"
    
    payload = {
        "requests": [{
            "image": {"content": image_content},
            "features": [{"type": "TEXT_DETECTION"}],
            "imageContext": {
                "languageHints": ["ar"]  # Hint pour l'arabe
            }
        }]
    }
    
    response = requests.post(url, json=payload)
    result = response.json()
    
    if "error" in result:
        raise Exception(f"Vision API Error: {result['error']}")
    
    # Extraire le texte
    try:
        text = result["responses"][0]["fullTextAnnotation"]["text"]
        return text.strip()
    except (KeyError, IndexError):
        return ""


def translate_with_gemini(arabic_text, api_key, model=GEMINI_MODEL):
    """Traduit du texte arabe vers le français avec Gemini"""
    
    if not arabic_text.strip():
        return ""
    
    url = f"https://generativelanguage.googleapis.com/v1beta/models/{model}:generateContent?key={api_key}"
    
    prompt = f"""Traduis ce texte arabe en français. 
Garde la mise en forme et les paragraphes. 
Ne rajoute aucun commentaire, juste la traduction.

Texte arabe:
{arabic_text}

Traduction française:"""
    
    payload = {
        "contents": [{
            "parts": [{"text": prompt}]
        }],
        "generationConfig": {
            "temperature": 0.3,
            "maxOutputTokens": 8192
        }
    }
    
    response = requests.post(url, json=payload)
    result = response.json()
    
    if "error" in result:
        raise Exception(f"Gemini API Error: {result['error']}")
    
    try:
        translation = result["candidates"][0]["content"]["parts"][0]["text"]
        return translation.strip()
    except (KeyError, IndexError):
        return "[Erreur de traduction]"


def process_book(pdf_path, api_key, output_prefix="livre"):
    """
    Pipeline complet: PDF → OCR → Traduction → Export
    """
    print("=" * 60)
    print("🚀 PIPELINE OCR ARABE + TRADUCTION FRANÇAIS")
    print("=" * 60)
    
    # 1. Convertir PDF en images
    image_paths = pdf_to_images(pdf_path)
    total_pages = len(image_paths)
    
    # 2. OCR + Traduction pour chaque page
    results = []
    
    for i, img_path in enumerate(image_paths, 1):
        print(f"\n📖 Page {i}/{total_pages}")
        
        # OCR
        print("  🔍 OCR en cours...")
        try:
            arabic_text = ocr_with_vision_api(img_path, api_key)
            print(f"  ✓ OCR OK ({len(arabic_text)} caractères)")
        except Exception as e:
            print(f"  ❌ Erreur OCR: {e}")
            arabic_text = ""
        
        # Traduction
        if arabic_text:
            print("  🌍 Traduction en cours...")
            try:
                french_text = translate_with_gemini(arabic_text, api_key)
                print(f"  ✓ Traduction OK ({len(french_text)} caractères)")
            except Exception as e:
                print(f"  ❌ Erreur traduction: {e}")
                french_text = "[Erreur]"
        else:
            french_text = ""
        
        results.append({
            "page": i,
            "arabic": arabic_text,
            "french": french_text
        })
        
        # Pause pour éviter rate limiting
        time.sleep(1)
    
    # 3. Export JSON
    json_path = f"{output_prefix}_traduction.json"
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(results, f, ensure_ascii=False, indent=2)
    print(f"\n✅ JSON exporté: {json_path}")
    
    # 4. Export Excel
    excel_path = f"{output_prefix}_traduction.xlsx"
    export_to_excel(results, excel_path)
    print(f"✅ Excel exporté: {excel_path}")
    
    return results, json_path, excel_path


def export_to_excel(results, excel_path):
    """Exporte les résultats en Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Traductions"
    
    # En-têtes
    headers = ["Page", "Texte Arabe", "Traduction Française"]
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Données
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row_idx, item in enumerate(results, 2):
        # Page
        ws.cell(row=row_idx, column=1, value=item["page"]).alignment = Alignment(horizontal="center")
        
        # Texte arabe (aligné à droite pour RTL)
        arabic_cell = ws.cell(row=row_idx, column=2, value=item["arabic"])
        arabic_cell.alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)
        
        # Traduction française
        french_cell = ws.cell(row=row_idx, column=3, value=item["french"])
        french_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        
        # Bordures
        for col in range(1, 4):
            ws.cell(row=row_idx, column=col).border = thin_border
    
    # Largeurs de colonnes
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 60
    
    # Hauteur des lignes (auto-ajustement approximatif)
    for row_idx in range(2, len(results) + 2):
        ws.row_dimensions[row_idx].height = 100
    
    wb.save(excel_path)


def main():
    """Point d'entrée principal"""
    import sys
    
    if len(sys.argv) < 2:
        print("""
Usage: python arabic_ocr_translate.py <fichier.pdf> [nom_sortie]

Exemple:
    python arabic_ocr_translate.py mon_livre.pdf livre1
    
Cela créera:
    - livre1_traduction.json
    - livre1_traduction.xlsx
        """)
        sys.exit(1)
    
    pdf_path = sys.argv[1]
    output_prefix = sys.argv[2] if len(sys.argv) > 2 else Path(pdf_path).stem
    
    if not os.path.exists(pdf_path):
        print(f"❌ Fichier non trouvé: {pdf_path}")
        sys.exit(1)
    
    # Installer les dépendances
    print("📦 Installation des dépendances...")
    install_dependencies()
    
    # Lancer le pipeline
    results, json_path, excel_path = process_book(pdf_path, API_KEY, output_prefix)
    
    print("\n" + "=" * 60)
    print("🎉 TERMINÉ!")
    print(f"   Pages traitées: {len(results)}")
    print(f"   JSON: {json_path}")
    print(f"   Excel: {excel_path}")
    print("=" * 60)


if __name__ == "__main__":
    main()
